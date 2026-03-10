# goldenpages_rubric_to_xlsx_dom.py
# -*- coding: utf-8 -*-
import re, time, sys, traceback
from urllib.parse import urljoin, urlparse, urlunparse, quote, parse_qsl, urlencode

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, StaleElementReferenceException, WebDriverException,
    NoSuchElementException, InvalidArgumentException
)
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

RUBRIC_URL = input("Enter the url of the category: ")
XLSX   = input("Name of xlsx: ")
OUT_XLSX = XLSX +".xlsx"
WAIT       = 15
HEADLESS   = True

# ----------------------- driver & utils -----------------------

def mk_driver():
    opts = webdriver.ChromeOptions()
    if HEADLESS: opts.add_argument("--headless=new")
    opts.add_argument("--window-size=1440,900")
    opts.add_argument("--no-sandbox"); opts.add_argument("--disable-gpu"); opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122 Safari/537.36")
    prefs = {"profile.managed_default_content_settings.images": 2, "profile.managed_default_content_settings.fonts": 2}
    opts.add_experimental_option("prefs", prefs)
    d = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    try:
        d.execute_cdp_cmd("Network.enable", {})
        d.execute_cdp_cmd("Network.setBlockedURLs", {"urls": ["*.jpg","*.jpeg","*.png","*.gif","*.webp","*.svg","*.woff","*.woff2","*.ttf","*.otf"]})
    except Exception:
        pass
    d.set_page_load_timeout(60)
    return d

def wait_ready(drv, t=WAIT):
    WebDriverWait(drv, t).until(lambda d: d.execute_script("return document.readyState") == "complete")

def norm_spaces(s): 
    return re.sub(r"\s+", " ", (s or "")).strip()

def norm_phones(items):
    out, seen = [], set()
    for t in items:
        t = (t or "").replace("\u00a0", " ")
        t = re.sub(r"[^\d\+\(\)\-\s]", "", t)
        t = re.sub(r"\s+", " ", t).strip()
        if not t: continue
        if not re.search(r"\d{2}[\s\-\)]?\d{2}", t):  # looks like a phone
            continue
        if t not in seen:
            seen.add(t); out.append(t)
    return out

def company_id_from_href(href):
    m = re.search(r"[?&]Id=(\d+)", href or "")
    return m.group(1) if m else ""

def normalize_url(u: str) -> str:
    """Trim & percent-encode path/query safely; keep reserved chars."""
    u = str(u or "").strip()
    if not re.match(r"^https?://", u, re.I):
        raise ValueError(f"Bad URL (missing scheme): {u!r}")
    parts = urlparse(u)
    path = quote(parts.path, safe="/%:@")
    query = urlencode(parse_qsl(parts.query, keep_blank_values=True), doseq=True)
    return urlunparse((parts.scheme, parts.netloc, path, parts.params, query, parts.fragment))

def safe_get(drv, url):
    """driver.get with normalization/retry to avoid InvalidArgumentException."""
    u = str(url).strip()
    try:
        drv.get(u); wait_ready(drv); return
    except InvalidArgumentException:
        pass
    except WebDriverException as e:
        if "invalid argument" not in str(e).lower(): raise
    u2 = normalize_url(u)
    print(f"[RETRY] normalized URL -> {u2}")
    drv.get(u2); wait_ready(drv)

# ----------------------- page helpers -----------------------

def find_company_cards(drv):
    WebDriverWait(drv, WAIT).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "section.gp_company"))
    )
    return drv.find_elements(By.CSS_SELECTOR, "section.gp_company")

def extract_name_and_href(card):
    a = card.find_element(By.CSS_SELECTOR, "h3.h3.mb-0 a[href*='/company/?Id=']")
    return (a.text or "").strip(), a.get_attribute("href") or ""

def extract_address_block(card):
    try:
        p = card.find_element(By.CSS_SELECTOR, "div.gp_wrap_address p")
    except NoSuchElementException:
        return "", "", ""
    # city
    city = ""
    try:
        city_a = p.find_element(By.CSS_SELECTOR, "a[href*='/city/?Id=']")
        city = norm_spaces(city_a.text)
    except NoSuchElementException:
        pass
    addr = norm_spaces(p.get_attribute("innerText"))
    # landmarks
    landmark = ""
    try:
        lm = card.find_element(By.XPATH, ".//div[contains(@class,'gp_wrap_address')]/following-sibling::div[contains(@class,'gp_job2')][1]//p")
        landmark = norm_spaces(lm.text)
    except Exception:
        pass
    return addr, city, landmark

def reveal_and_collect_phones(card, org_id, drv):
    phones = []
    if not org_id: return phones
    sel_container = f"div#PhonesByOrg_{org_id}"
    try:
        container = card.find_element(By.CSS_SELECTOR, sel_container)
    except NoSuchElementException:
        return phones

    def current_phones():
        tel_links = [a.get_attribute("href") or "" for a in container.find_elements(By.CSS_SELECTOR, "a[href^='tel:']")]
        from_links = [re.sub(r"^tel:", "", t) for t in tel_links if t]
        text = container.text or ""
        patt = r"(?:\+?998[\s\-]?\(?\d{2}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}|\(\d{2}\)\s*\d{3}[\s\-]?\d{2}[\s\-]?\d{2})"
        from_text = re.findall(patt, text)
        return norm_phones(from_links + from_text)

    got = current_phones()
    if got: return got

    for attempt in range(3):
        try:
            btn = container.find_element(By.CSS_SELECTOR, "button.gp_btn_s1qp, button[onclick*=\"'phones'\"]")
            drv.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            try: btn.click()
            except Exception: drv.execute_script("arguments[0].click();", btn)
        except NoSuchElementException:
            # try clicking masked anchor
            try:
                masked = container.find_element(By.CSS_SELECTOR, "a[href='javascript:void(0)']")
                drv.execute_script("arguments[0].scrollIntoView({block:'center'});", masked)
                try: masked.click()
                except Exception: drv.execute_script("arguments[0].click();", masked)
            except Exception:
                pass

        try:
            WebDriverWait(container, 4).until(
                lambda _:
                    container.find_elements(By.CSS_SELECTOR, "a[href^='tel:']") or
                    re.search(r"\d{2}[\s\-\)]?\d{2}", container.text or "")
            )
        except TimeoutException:
            pass

        got = current_phones()
        if got:
            phones = got
            break
        time.sleep(0.6)
    return phones

def try_click_next_page(drv):
    try:
        nav = drv.find_element(By.CSS_SELECTOR, "nav.gp_navigation")
    except NoSuchElementException:
        return False
    try:
        nxt_li = nav.find_element(By.CSS_SELECTOR, "li.gp_next")
    except NoSuchElementException:
        return False
    if "disabled" in (nxt_li.get_attribute("class") or ""):
        return False
    try:
        marker = drv.find_element(By.CSS_SELECTOR, "section.gp_company h3.h3.mb-0")
    except NoSuchElementException:
        marker = None
    try:
        el = nxt_li.find_element(By.CSS_SELECTOR, "a, span, *")
    except NoSuchElementException:
        el = nxt_li
    try:
        drv.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        el.click()
    except Exception:
        drv.execute_script("arguments[0].click();", el)
    if marker:
        try:
            WebDriverWait(drv, WAIT).until(EC.staleness_of(marker))
        except TimeoutException:
            pass
    try:
        WebDriverWait(drv, WAIT).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "section.gp_company"))
        )
        return True
    except TimeoutException:
        return False

# ----------------------- scraping flow -----------------------

def scrape_rubric(url):
    drv = mk_driver()
    rows = []
    try:
        print(f"[OPEN] {url!r}")
        safe_get(drv, url)  # robust GET with normalization/retry

        page = 1
        seen_company_ids = set()
        while True:
            print(f"[PAGE {page}] collect companies…")
            cards = find_company_cards(drv)
            print(f"  └─ found {len(cards)} cards")

            for idx, card in enumerate(cards, 1):
                try:
                    name, href = extract_name_and_href(card)
                    abs_href = urljoin(url, href)
                    org_id = company_id_from_href(abs_href)

                    addr, city, landmark = extract_address_block(card)
                    phones = reveal_and_collect_phones(card, org_id, drv)

                    if org_id and org_id in seen_company_ids:
                        continue
                    if org_id:
                        seen_company_ids.add(org_id)

                    rows.append({
                        "Наименования компании": name,
                        "Телефон номер": "\n".join(phones),
                        "Город": city,
                        "Адрес": addr,
                        "Ориентир": landmark,
                        "Email": "",
                        "Ссылка": abs_href
                    })

                    print(f"    [{idx}/{len(cards)}] {name[:60]} | phones: {', '.join(phones) if phones else '—'}")
                except Exception as e:
                    print(f"    [{idx}/{len(cards)}] ERROR: {e}")
                    traceback.print_exc()
                time.sleep(0.05)

            if not try_click_next_page(drv):
                break
            page += 1

        print(f"[DONE] total companies: {len(rows)}")
        return rows
    finally:
        try: drv.quit()
        except: pass

# ----------------------- excel writer -----------------------

def write_excel(category_title, category_url, rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"

    ws["C3"] = "Направления"; ws["C3"].font = Font(bold=True, color="FF0000")
    ws["C4"] = category_title
    ws["D3"] = "Ссылка"; ws["D3"].font = Font(bold=True, color="FF0000")
    ws["D4"] = category_url
    ws["E3"] = "Кол-во номеров"; ws["E3"].font = Font(bold=True, color="FF0000")
    ws["E4"] = f"Найдено организаций: {len(rows)}"

    headers = ["№","Наименования компании","Телефон номер","Город","Адрес","Ориентир","Email","Ссылка"]
    start = 7
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, 1):
        ws.cell(row=start+i, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=start+i, column=2, value=r.get("Наименования компании",""))
        ws.cell(row=start+i, column=3, value=r.get("Телефон номер",""))
        ws.cell(row=start+i, column=4, value=r.get("Город",""))
        ws.cell(row=start+i, column=5, value=r.get("Адрес",""))
        ws.cell(row=start+i, column=6, value=r.get("Ориентир",""))
        ws.cell(row=start+i, column=7, value=r.get("Email",""))
        u = r.get("Ссылка","")
        ws.cell(row=start+i, column=8, value=u)
        if u:
            ws.cell(row=start+i, column=8).hyperlink = u
            ws.cell(row=start+i, column=8).style = "Hyperlink"

    widths = {1:6, 2:42, 3:26, 4:18, 5:60, 6:42, 7:30, 8:45}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in ws.iter_rows(min_row=start+1, min_col=2, max_col=8, max_row=start+len(rows)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = f"A{start+1}"
    ws.auto_filter.ref = f"A{start}:H{start+len(rows)}"

    wb.save(out_path)
    print(f"[SAVED] {out_path}")

def _pick_cli_args():
    """Return (url, out) from sys.argv while ignoring Jupyter flags like -f."""
    url, out = RUBRIC_URL, OUT_XLSX
    for a in sys.argv[1:]:
        a = str(a).strip()
        if not a:
            continue
        if a.startswith("--url="):
            url = a.split("=", 1)[1].strip()
        elif a.startswith("--out="):
            out = a.split("=", 1)[1].strip()
        elif re.match(r"^https?://", a, re.I):
            url = a
        elif a.lower().endswith(".xlsx"):
            out = a
        else:
            # ignore unknown flags/args (e.g., -f, connection files, etc.)
            continue
    return url, out

# ----------------------- main -----------------------


def main():
    url, out = _pick_cli_args()
    # safety: if something weird still slipped in, fall back to default
    if not re.match(r"^https?://", url, re.I):
        print(f"[WARN] Bad or missing URL '{url}', falling back to default.")
        url = RUBRIC_URL
    title = "GoldenPages Rubric " + (urlparse(url).query or url)
    rows = scrape_rubric(url)
    write_excel(title, url, rows, out)

if __name__ == "__main__":
    main()
